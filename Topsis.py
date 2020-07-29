import xlwt
import openpyxl

from openpyxl import load_workbook
wb = load_workbook('D:/Python project/BigCode/singleQue.xlsx')
ws2 = wb['查找算法TOPSIS'] # 查找算法
ws3 = wb['排序算法TOPSIS'] # 排序算法
ws4 = wb['树结构TOPSIS'] # 树结构
ws5 = wb['数字操作TOPSIS']# 数字操作
ws6 = wb['数组TOPSIS']# 数组
ws7 = wb['图结构TOPSIS']# 图结构
ws8 = wb['线性表TOPSIS']# 线性表
ws9 = wb['字符串TOPSIS']# 字符串

ws = ws9 # 做另外题目类型的处理修改此行即刻
allData =[]
for i in ws.rows:
    newData=[]
    for j in i:
        newData.append(j.value)
    allData.append(newData)

allData1=[]
max =0
for i in range(1,len(allData)):
    newData=[]
    newData.append(allData[i][0])
    newData.append(allData[i][3])
    newData.append(allData[i][2]/allData[i][1])
    if (allData[i][2]/allData[i][1]>max):
        max = allData[i][2]/allData[i][1]
    allData1.append(newData)

for i in range(0,len(allData1)):
    allData1[i][2] = max-allData1[i][2]


total=0

for i in range(0,len(allData1)):
    total=total+allData1[i][1]**2+allData1[i][2]**2
total=total**0.5

max1=0
max2=0
min1=1
min2=1

for i in range(0,len(allData1)):
    allData1[i][1] = allData1[i][1]/total
    allData1[i][2] = allData1[i][2] / total
    if allData1[i][1]>max1:
        max1=allData1[i][1]
    if allData1[i][1]<min1:
        min1=allData1[i][1]
    if allData1[i][2]>max2:
        max2=allData1[i][2]
    if allData1[i][2]<min2:
        min2=allData1[i][2]


allDataD=[]
for i in range(0,len(allData1)):
    newData=[]
    newData.append(allData1[i][0])
    Dp=((max1-allData1[i][1])**2+(max2-allData1[i][2])**2)**0.5
    Dm=((min1-allData1[i][1])**2+(min2-allData1[i][2])**2)**0.5
    newData.append(Dp)
    newData.append(Dm)
    newData.append(Dm/(Dp+Dm))
    allDataD.append(newData)

for i in range(0,len(allData1)):
    ws9.cell(2+i,5).value=allDataD[i][3]

print(max1)
print(min1)
print(max2)
print(min2)

print(total)
wb.save('D:/Python project/BigCode/singleQue.xlsx')
print(allDataD)

