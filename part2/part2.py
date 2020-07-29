import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np


def export(filename,data):
    wb = openpyxl.Workbook()
    ws = wb.active
    for d in data:
        ws.append(d)
    wb.save(filename)


dic = {}
dic_Qnum = {}
l = []
l2 = []
wb1 = openpyxl.load_workbook("D:/Python project/BigCode/part2/data/que.xlsx")
ws1 = wb1.active
wb2 = openpyxl.load_workbook("D:/Python project/BigCode/part2/data/personalAll.xlsx")
ws2 = wb2.active
for row in ws1.iter_rows(min_row=2,values_only=True):
    l.append(row)
for row in ws2.iter_rows(min_row=2,values_only=True):
    l2.append(row)

for data in l:
    if dic.get(data[0])==None:
        dic[data[0]] = data[6]
        dic_Qnum[data[0]] = data[5]
    else:
        dic[data[0]] = dic.get(data[0]) + data[6]
        dic_Qnum[data[0]] = dic_Qnum.get(data[0]) + data[5]
data = [
    ["Type","percentage"],
    ["图结构",round(dic.get("图结构")*1.0/dic_Qnum.get("图结构"),2)],
    ["字符串",round(dic.get("字符串")*1.0/dic_Qnum.get("字符串"),2)],
    ["排序算法",round(dic.get("排序算法")*1.0/dic_Qnum.get("排序算法"),2)],
    ["数字操作",round(dic.get("数字操作")*1.0/dic_Qnum.get("数字操作"),2)],
    ["数组",round(dic.get("数组")*1.0/dic_Qnum.get("数组"),2)],
    ["查找算法",round(dic.get("查找算法")*1.0/dic_Qnum.get("查找算法"),2)],
    ["树结构",round(dic.get("树结构")*1.0/dic_Qnum.get("树结构"),2)],
    ["线性表",round(dic.get("线性表")*1.0/dic_Qnum.get("线性表"),2)]
]
data1 = [
    ["Type","face_num","upload_num"],
    ["图结构",dic.get("图结构"),dic_Qnum.get("图结构")],
    ["字符串",dic.get("字符串"),dic_Qnum.get("字符串")],
    ["排序算法",dic.get("排序算法"),dic_Qnum.get("排序算法")],
    ["数字操作",dic.get("数字操作"),dic_Qnum.get("数字操作")],
    ["数组",dic.get("数组"),dic_Qnum.get("数组")],
    ["查找算法",dic.get("查找算法"),dic_Qnum.get("查找算法")],
    ["树结构",dic.get("树结构"),dic_Qnum.get("树结构")],
    ["线性表",dic.get("线性表"),dic_Qnum.get("线性表")]
]

export("Percentage.xlsx",data)
export("Num.xlsx",data1)
percent = pd.read_excel("Percentage.xlsx")
number = pd.read_excel("Num.xlsx")
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False
plt.style.use("ggplot")
plt.bar(x = range(percent.shape[0]),
        height= percent.percentage,
        tick_label = percent.Type,
        color = 'steelblue')
plt.ylabel("面向用例占比(%)")
for x,y in enumerate(percent.percentage):
    plt.text(x,y+0.005,y,ha='center')
plt.show()
plt.close()

Type = number.Type              #类型
face_sample = number.face_num   #面向用例
total_num = number.upload_num   #总提交次数

bar_width = 0.4
plt.bar(x = np.arange(len(Type)),
        height= total_num,
        label = "总提交次数",
        color = 'steelblue',
        width = bar_width)
plt.bar(x=np.arange(len(Type))+bar_width,
        height=face_sample,
        label = "面向用例提交次数",
        color = "indianred",
        width= bar_width)
plt.xticks(np.arange(len(Type))+0.2,Type)
plt.ylabel("提交次数")
li = list(enumerate(percent.percentage))
count = 0
for x,y in enumerate(number.upload_num):
    plt.text(x+0.2,y+800,str(li[count][1])+'%',ha='center')
    plt.text(x,y+200,y,ha='center')
    count +=1
for x,y in enumerate(number.face_num):
    plt.text(x+0.4,y+40,y,ha='center')
plt.legend()
plt.show()
plt.close()

#根据难度系数进行排序
data2 = []
#类型:[答题人数，原始分数]
dic1={
    "图结构":[0,0],
    "字符串":[0,0],
    "排序算法":[0,0],
    "数字操作":[0,0],
    "数组":[0,0],
    "查找算法":[0,0],
    "树结构":[0,0],
    "线性表":[0,0]
}
#类型:[答题人数，实际分数]
dic2={
    "图结构":[0,0],
    "字符串":[0,0],
    "排序算法":[0,0],
    "数字操作":[0,0],
    "数组":[0,0],
    "查找算法":[0,0],
    "树结构":[0,0],
    "线性表":[0,0]
}

for d in l:
    dic1.get(d[0])[0] += d[2]
    dic1.get(d[0])[1] += d[8]
    dic2.get(d[0])[0] += d[2]
    dic2.get(d[0])[1] += d[10]

#L=1-X/W 难度系数定义为原始样本均分/总分
nanduxishu = []

for item in dic1:
    nanduxishu.append([item,1-dic1.get(item)[1]/dic1.get(item)[0]/100.0])
res = sorted(nanduxishu,key=lambda x:x[1],reverse=True)
res.append(["Type","xishu"])
res.reverse()
export("难度系数.xlsx",res)

ndxs = pd.read_excel("难度系数.xlsx")
plt.style.use("ggplot")
plt.bar(x = range(ndxs.shape[0]),
        height= ndxs.xishu,
        tick_label = ndxs.Type,
        color = 'steelblue')
plt.ylabel("难度系数(数值越大表明越难)")
for x,y in enumerate(ndxs.xishu):
    plt.text(x,y+0.0005,round(y,4),ha='center')
plt.show()
plt.close()

#自定义难度系数：L=（1-实际平均分/试卷分数）*0.5 + face_percentage*0.3 + （1-原始平均分/试卷分数）*0.2
nanduxishu_sd = []
wb_percent = openpyxl.load_workbook("Percentage.xlsx")
ws_percent = wb_percent.active
dic_percent={
    "图结构":0,
    "字符串":0,
    "排序算法":0,
    "数字操作":0,
    "数组":0,
    "查找算法":0,
    "树结构":0,
    "线性表":0
}
for row in ws_percent.iter_rows(min_row=2,values_only=True):
    dic_percent[row[0]] = row[1]
for item in dic2:
    nanduxishu_sd.append([item,(1-dic2.get(item)[1]/dic1.get(item)[0]/100.0)*0.5
                       + dic_percent.get(item)*0.3
                       + (1-dic1.get(item)[1]/dic1.get(item)[0]/100.0)*0.2])
res = sorted(nanduxishu_sd,key=lambda x:x[1],reverse=True)
res.append(["Type","xishu"])
res.reverse()
export("难度系数_sd.xlsx",res)

ndxs_sd = pd.read_excel("难度系数_sd.xlsx")
plt.style.use("ggplot")
plt.bar(x = range(ndxs_sd.shape[0]),
        height= ndxs_sd.xishu,
        tick_label = ndxs_sd.Type,
        color = 'steelblue')
plt.ylabel("难度系数(数值越大表明越难)")
for x,y in enumerate(ndxs_sd.xishu):
    plt.text(x,y+0.008,round(y,4),ha='center')
plt.show()
plt.close()